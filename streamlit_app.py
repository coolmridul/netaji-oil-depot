import streamlit as st
import pandas as pd
import numpy as np
from app import *
import requests
import matplotlib.pyplot as plt
from send_whatsapp import *
from openpyxl import load_workbook

st.set_page_config(layout="wide")
hide_github_icon = """
<style>
#MainMenu {
  visibility: hidden;
}
<style>
"""
st.markdown(hide_github_icon, unsafe_allow_html=True)


st.title("PAYMENT TAKADA")

df1=pd.read_excel('list.xlsx')

def call_list():
    df_final=pd.read_excel('final.xlsx')
    df2=df_final[df_final['is_done'] == 0].reset_index(drop=True)

    return df_final,df2

df_final,df2 = call_list()

col1, col2, col3 = st.columns(3)

with col1:
    option = st.selectbox(
        "Broker",df1['Broker'].to_list(),
    )
with col2:
    option1 = st.text_input("Party Name")
    # option1 = st.selectbox(
    #     "Party",("D K ", "Home phone", "Mobile phone"),
    # )
with col3:
    option2 = st.text_input("Amount: ")

col6, col7, col8 = st.columns(3)

with col6:
    option4 = st.date_input("Invoice Date",format="DD/MM/YYYY")
with col7:
    option5 = st.text_input("Invoice Number")
with col8:
    option9 = st.selectbox(
        "Company",
    ("NODPL","NODGDM","NODKOL","SSCO")
    )

df9 = df1[df1['Broker'] == option].reset_index()

st.write("Email: ")
st.code(df9['Email'].iloc[0], language="markdown")
st.write("Number: ")
st.code(df9['Number'].iloc[0], language="markdown")

if st.button("Submit",type="primary"):
    data = pd.DataFrame([{"ID":df_final['ID'].max()+1,"Broker":option,"Party":option1,"Amount":option2,"IDate":option4,"INumber":option5,"is_done":0,"Company":option9}])
    with pd.ExcelWriter('final.xlsx', engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:
        data.to_excel(writer, sheet_name='Sheet1', index=False, header=None, startrow= writer.sheets['Sheet1'].max_row)
    
    df_final,df2=call_list()



st.header("PENDING LIST")

df2['is_widget'] = False

edited_df = st.data_editor(df2,disabled=("ID","Broker","Party","Amount","IDate","INumber","is_done","Company"),
                           column_order=("Broker","Company","Party","Amount","IDate","INumber","is_widget"),use_container_width=True)

col4, col5,col7,col14,col15 = st.columns(5)

with col5:
    if st.button("Send SMS",type="primary"):
        edited_df2=edited_df[edited_df['is_widget'] == True].reset_index()
        edited_df2['INumber'] = edited_df2['INumber'].fillna(0)
        edited_df2['INumber'] = edited_df2['INumber'].astype(int)
        edited_df2['Amount'] = edited_df2['Amount'].astype(int)
        if edited_df2.shape[0]:
            for j in edited_df2['Broker'].unique().tolist():
                edited_df3 = edited_df2[edited_df2['Broker'] == j].reset_index()
                df11 = df1[df1['Broker'] == j].reset_index()
                df11['GroupID'] = df11['GroupID'].fillna('')
                if df11['GroupID'].iloc[0] != '':
                    if edited_df3.shape[0]:
                        final_text_message = 'PLEASE CLEAR THE DUES \n\n'
                        for l in range(0,edited_df3.shape[0]):
                            if edited_df3['INumber'].astype(str).iloc[l] == 'nan' or edited_df3['INumber'].astype(str).iloc[l] == '0':
                                final_text_message += edited_df3['Company'].astype(str).iloc[l]+ " - " + edited_df3['Party'].iloc[l] +  "  "+ str(edited_df3['Amount'].iloc[l]) +"/- \n"
                            else:
                                final_text_message += edited_df3['Company'].astype(str).iloc[l]+ " - " + edited_df3['Party'].iloc[l] + "  INV-"+edited_df3['INumber'].astype(str).iloc[l]+"  AMT-" +str(edited_df3['Amount'].iloc[l]) +"/-  Dt- " + edited_df3['IDate'].dt.strftime('%d-%m-%Y').iloc[l] + "\n\n"
                        
                        send_payment_whatsapp(final_text_message,df11['GroupID'].iloc[0])
                        st.toast('Whatsapp Message Sent' , icon="✅")
                else:
                    st.toast('Whatsapp Group not available for '+j, icon="⚠️")
        else:
            st.toast('Please select from the list' , icon="⚠️")
with col4:
    if st.button("Send Email",type="primary"):
        edited_df1=edited_df[edited_df['is_widget'] == True].reset_index()
        df1['Email'] = df1['Email'].fillna('')
        if edited_df1.shape[0]:
            for i in range(0,edited_df1.shape[0]):
                df10 = df1[df1['Broker'] == edited_df1['Broker'].iloc[i]].reset_index()
                text_invoice = "" if edited_df1['IDate'].dt.strftime('%Y-%m-%d').iloc[i] == '' else "INV - <b>"+edited_df1['INumber'].astype(str).iloc[i] + "</b> Dt - <b>" + edited_df1['IDate'].dt.strftime('%d-%m-%Y').iloc[i] + "</b>"
                if df10['Email'].iloc[0] == '' or df10['Email'].iloc[0] == 'nan':
                     st.toast('Email Not Found' , icon="❌")
                else:
                    send_email(subject,edited_df1['Amount'].iloc[i],edited_df1['Party'].iloc[i],text_invoice, sender, [df10['Email'].iloc[0]], password)
                    st.toast('Email Sent' , icon="✅")
        else:
            st.toast('Please select from the list' , icon="⚠️")

with col7:
    if st.button('Delete',type="primary"):
        edited_df1=edited_df[edited_df['is_widget'] == True].reset_index()
        if edited_df1.shape[0]:
            workbook = load_workbook('final.xlsx')
            sheet = workbook['Sheet1']
            for i in range(0,edited_df1.shape[0]):
                sheet.cell(row=edited_df1['ID'].iloc[i]+2, column=7, value=1)
            workbook.save('final.xlsx')
            # df_final,df2=call_list()
            st.rerun()
        else:
            st.toast('Please select from the list above', icon="⚠️")



with col14:
    if st.button('SEND SMS ALL',type="primary"):
        edited_df5=edited_df.copy()
        edited_df5['INumber'] = edited_df5['INumber'].fillna(0)
        edited_df5['INumber'] = edited_df5['INumber'].astype(int)
        edited_df5['Amount'] = edited_df5['Amount'].astype(int)
        if edited_df5.shape[0]:
            for j in edited_df5['Broker'].unique().tolist():
                edited_df3 = edited_df5[edited_df5['Broker'] == j].reset_index()
                df11 = df1[df1['Broker'] == j].reset_index()
                df11['GroupID'] = df11['GroupID'].fillna('')
                if df11['GroupID'].iloc[0] != '':
                    if edited_df3.shape[0]:
                        final_text_message = 'PLEASE CLEAR THE DUES \n\n'
                        for l in range(0,edited_df3.shape[0]):
                            if edited_df3['INumber'].astype(str).iloc[l] == 'nan' or edited_df3['INumber'].astype(str).iloc[l] == '0':
                                final_text_message += edited_df3['Company'].astype(str).iloc[l]+ " - " + edited_df3['Party'].iloc[l] +  "  "+ str(edited_df3['Amount'].iloc[l]) +"/- \n"
                            else:
                                final_text_message += edited_df3['Company'].astype(str).iloc[l]+ " - " + edited_df3['Party'].iloc[l] + "  INV-"+edited_df3['INumber'].astype(str).iloc[l]+"  AMT-" +str(edited_df3['Amount'].iloc[l]) +"/-  Dt- " + edited_df3['IDate'].dt.strftime('%d-%m-%Y').iloc[l] + "\n\n"
                        
                        send_payment_whatsapp(final_text_message,df11['GroupID'].iloc[0])
                        st.toast('Whatsapp Message Sent to '+j , icon="✅")
                else:
                    st.toast('Whatsapp Group not available for '+j, icon="⚠️")


with col15:
    if st.button('SEND EMAIL ALL',type="primary"):
        edited_df6=edited_df.copy()
        df1['Email'] = df1['Email'].fillna('')
        if edited_df6.shape[0]:
            for i in range(0,edited_df6.shape[0]):
                df10 = df1[df1['Broker'] == edited_df6['Broker'].iloc[i]].reset_index()
                text_invoice = "" if edited_df6['IDate'].dt.strftime('%Y-%m-%d').iloc[i] == '' else "INV - <b>"+edited_df6['INumber'].astype(str).iloc[i] + "</b> Dt - <b>" + edited_df6['IDate'].dt.strftime('%d-%m-%Y').iloc[i] + "</b>"
                if df10['Email'].iloc[0] == '' or df10['Email'].iloc[0] == 'nan':
                     st.toast('Email Not Found' , icon="❌")
                else:
                    send_email(subject,edited_df6['Amount'].iloc[i],edited_df6['Party'].iloc[i],text_invoice, sender, [df10['Email'].iloc[0]], password)
                    st.toast('Email Sent to '+ edited_df6['Broker'].iloc[i], icon="✅")


st.header("ANALYSIS")

@st.cache_data
def convert_df(df):
    return df.to_csv().encode("utf-8")

csv = convert_df(df_final)

st.download_button(
    label="Download data as CSV",
    data=csv,
    file_name="final.csv",
    mime="text/csv",
)

col10, col11 = st.columns(2)
with col10:
    df2['Amount'] = df2['Amount'].astype('int')
    df20 = df2.groupby('Broker')['Amount'].sum().reset_index().sort_values('Amount',ascending=False).reset_index(drop=True)
    st.dataframe(df20,use_container_width=True)
with col11:
    df2['Amount'] = df2['Amount'].astype('int')
    df20 = df2.groupby('Company')['Amount'].sum().reset_index().sort_values('Amount',ascending=False).reset_index(drop=True)
    st.dataframe(df20,use_container_width=True)

col12, col13 = st.columns(2)
with col12:
    df2['Amount'] = df2['Amount'].astype('int')
    df22 = df2.groupby('Party')['Amount'].sum().reset_index().sort_values('Amount',ascending=False).reset_index(drop=True)
    st.dataframe(df22,use_container_width=True)




# df20['test'] = df20['Broker'] + " : " + df20['Amount'].astype(str)
# # Pie chart, where the slices will be ordered and plotted counter-clockwise:
# labels = df20['Broker'].tolist()
# sizes = df20['Amount'].tolist()
# new_lendgent = df20['test'].tolist()


# fig1, ax1 = plt.subplots()
# ax1.pie(sizes, labels=labels, startangle=180)
# ax1.legend(new_lendgent, loc='lower left', bbox_to_anchor=(-0.1, 1.),fontsize=8)

# st.pyplot(fig1)
